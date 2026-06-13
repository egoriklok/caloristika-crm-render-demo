from __future__ import annotations

import json
import sqlite3
import time
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from html import escape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs"
DATA_DIR = ROOT / "data"
LOG_DIR = ROOT / "logs"
DB_PATH = DATA_DIR / "lunch_up_crm.sqlite"
TODAY = "2026-06-03"

PRODUCTION_ADDRESS = "Санкт-Петербург, Уральская улица, 13"
WALK_RADIUS_MIN = 40
WALK_RADIUS_KM = 3.2


def maps_url(platform: str, name: str, address: str) -> str:
    query = urllib.parse.quote(f"{name} {address} Санкт-Петербург")
    if platform == "2gis":
        return f"https://2gis.ru/spb/search/{query}"
    return f"https://yandex.ru/maps/2/saint-petersburg/search/{query}/"


def pb_url(query: str) -> str:
    return f"https://pb.nalog.ru/search.html?query={urllib.parse.quote(query)}"


def lead(
    name: str,
    segment: str,
    address: str,
    walk_min: int,
    score: int,
    fit_reason: str,
    offer: str,
    next_action: str,
    source_2gis: str | None = None,
    website: str | None = None,
    phone: str | None = None,
    email: str | None = None,
    nalog_query: str | None = None,
    expected_ogrn: str | None = None,
    notes: str = "",
) -> dict:
    source_2gis = source_2gis or maps_url("2gis", name, address)
    nalog_query = nalog_query or name
    priority = "A" if score >= 85 else "B" if score >= 72 else "C"
    return {
        "name": name,
        "segment": segment,
        "address": address,
        "walk_min": walk_min,
        "distance_band": "0-10 мин" if walk_min <= 10 else "11-25 мин" if walk_min <= 25 else "26-40 мин",
        "priority": priority,
        "score": score,
        "fit_reason": fit_reason,
        "offer": offer,
        "next_action": next_action,
        "phone": phone,
        "email": email,
        "website": website,
        "source_2gis": source_2gis,
        "source_yandex": maps_url("yandex", name, address),
        "pb_nalog_url": pb_url(nalog_query),
        "nalog_query": nalog_query,
        "expected_ogrn": expected_ogrn,
        "source_basis": "2GIS/Yandex Maps open listing + FNS public check",
        "notes": notes,
    }


OFFICE_OFFER = "Пилотная витрина Lunch Up для арендаторов: сэндвичи, роллы, салаты, десерты, 2 поставки в неделю."
COFFEE_OFFER = "Допродажа еды к кофе без кухни: 4-6 SKU сэндвичей/десертов на пробную неделю."
RETAIL_OFFER = "Локальная fresh-полка: сэндвичи, салаты, завтраки и десерты с тестом продаж по 10-14 SKU."
HOTEL_OFFER = "Снэк-бокс для лобби/апарт-отеля: завтраки, роллы, десерты и сэндвичи с регулярной поставкой."
MEDICAL_OFFER = "Перекусы для персонала и посетителей: аккуратная витрина с понятным сроком годности."
COWORKING_OFFER = "Холодильник/полка для резидентов: готовая еда на будни, учет списаний и повторный заказ."


PROSPECTS = [
    lead("БЦ Алмаз", "Бизнес-центр", "Уральская улица, 13 лит К", 1, 96, "Соседнее здание у производства; в 2ГИС указаны 3 этажа и 12 организаций в здании.", OFFICE_OFFER, "Позвонить в администрацию БЦ и предложить тестовую витрину для арендаторов.", "https://2gis.ru/spb/firm/5348552838583842", nalog_query="Алмаз бизнес центр Уральская 13"),
    lead("Урал Плаза", "Бизнес-центр", "Уральская улица, 19", 7, 94, "БЦ рядом с производством; в 2ГИС указано 27 организаций в здании.", OFFICE_OFFER, "Выйти на администрацию и запросить список арендаторов/оператора питания.", "https://2gis.ru/spb/firm/70000001045044337", website="https://ural-plaza.ru", nalog_query="Урал Плаза"),
    lead("Мануфактура", "Бизнес-центр", "Уральская улица, 17 к1 лит Д", 5, 88, "БЦ в пешей зоне производства; 2ГИС показывает 5 этажей и список организаций.", OFFICE_OFFER, "Предложить стартовую полку для арендаторов и доставку в один адрес.", "https://2gis.ru/spb/firm/5348552838650998", nalog_query="Мануфактура бизнес центр Уральская 17"),
    lead("Урок", "Бизнес-центр", "Уральская улица, 1 к2", 10, 93, "БЦ с высокой плотностью офисных арендаторов; в 2ГИС указано 37 организаций.", OFFICE_OFFER, "Начать с администрации, затем сделать обход арендаторов первого этажа.", "https://2gis.ru/spb/firm/5348552838484493", nalog_query="Урок бизнес центр Уральская 1"),
    lead("Baltis Plaza", "Бизнес-центр", "Средний проспект В.О., 88", 25, 91, "Крупный БЦ; 2ГИС показывает 9 этажей, 60 организаций и внутренние food-точки.", OFFICE_OFFER, "Предложить управляющей компании тест Lunch Up для арендаторов и переговорить с food-точками.", "https://2gis.ru/spb/firm/5348552838657634", website="https://baltisplaza.ru", nalog_query="Baltis Plaza"),
    lead("Биржа", "Бизнес-центр", "26-я линия В.О., 15 к2 лит А", 32, 88, "Крупный БЦ; в 2ГИС указано 16 этажей и 36 организаций в здании.", OFFICE_OFFER, "Связаться с администрацией и проверить наличие микромаркета/столовой.", "https://2gis.ru/firm/5348552838778777", nalog_query="Биржа бизнес центр 26 линия"),
    lead("Asgard", "Бизнес-центр", "Наличная улица, 44", 24, 82, "БЦ в 7 минутах от метро Приморская по профильным открытым карточкам.", OFFICE_OFFER, "Предложить регулярную поставку для офисов и переговорных.", "https://2gis.ru/spb/firm/70000001056226714", nalog_query="Asgard Наличная 44"),
    lead("Соверен", "Бизнес-центр", "Малый В.О. проспект, 22", 27, 90, "В 2ГИС указано 19 организаций в здании, включая супермаркет, банки и офисы.", OFFICE_OFFER, "Начать с управляющей компании и магазина Жизньмарт как якорной точки.", "https://2gis.ru/spb/firm/5348552838528578", nalog_query="Соверен бизнес центр"),
    lead("Gustaf", "Бизнес-центр", "Средний проспект В.О., 38 к1 лит К", 30, 84, "БЦ класса B+; открытые карточки указывают 7 этажей и 10 организаций.", OFFICE_OFFER, "Предложить полку для арендаторов и отдельный B2B-набор для клиник/салонов.", "https://2gis.ru/spb/geo/5348720342208864", nalog_query="Gustaf бизнес центр"),
    lead("БЦ Средний-4", "Бизнес-центр", "Средний проспект В.О., 4", 39, 74, "Деловой объект на границе 40-минутной пешей зоны.", OFFICE_OFFER, "Проверить через карты текущих арендаторов и администрацию.", nalog_query="Средний 4 бизнес центр"),
    lead("Сенатор, 17-я линия В.О.", "Бизнес-центр", "17-я линия В.О., 22", 37, 81, "Сеть БЦ в центре Васильевского острова; подходит для офисного питания.", OFFICE_OFFER, "Писать в отдел аренды/управляющую компанию, просить контакт facility manager.", "https://senator.spb.ru/contacts/", phone="+7 (812) 332-30-00", email="sales@senator.spb.ru", nalog_query="Сенатор бизнес центр Санкт-Петербург"),
    lead("БЦ 23-я линия, 2", "Бизнес-центр", "23-я линия В.О., 2", 30, 78, "Офисный объект на 23-й линии рядом с БЦ Биржа.", OFFICE_OFFER, "Проверить арендаторов по 2ГИС и предложить пробную доставку в один адрес.", nalog_query="23-я линия 2 бизнес центр"),
    lead("ТОЦ Остров", "Торгово-офисный центр", "Средний проспект В.О., 36", 31, 76, "Офисно-торговый формат на Среднем проспекте В.О.; релевантен для точек питания и офисов.", OFFICE_OFFER, "Собрать арендаторов food/office и сделать короткий маршрут обхода.", nalog_query="ТОЦ Остров Средний проспект"),
    lead("Смола", "Кофейня", "Уральская улица, 2 ст1", 8, 83, "Кофейня рядом с производством; 2ГИС показывает меню с выпечкой, десертами и завтраками.", COFFEE_OFFER, "Лично занести дегустационный набор: 2 сэндвича, 2 десерта, сырники.", "https://2gis.ru/spb/firm/70000001038555616", website="https://cafesmola.tilda.ws", nalog_query="Смола кофейня Уральская"),
    lead("Baggins Coffee, Уральская", "Кофейня", "Уральская улица, 21 лит А", 8, 86, "Сетевая кофейня в ЖК; 2ГИС содержит сайт, email и ОГРН рекламодателя.", COFFEE_OFFER, "Написать на общий email и параллельно зайти в точку с дегустацией.", "https://2gis.ru/spb/firm/70000001065317738", website="https://bagginscoffee.com", phone="8 (800) 600-70-15", email="info@bagginscoffee.ru", nalog_query="1177847162063", expected_ogrn="1177847162063"),
    lead("Молния", "Кафе/кофейня", "Уральская улица, 6 ст1", 6, 76, "Локальная точка питания рядом с производством и Уральской улицей.", COFFEE_OFFER, "Проверить формат кухни и предложить SKU для витрины/перекуса.", nalog_query="Молния кафе Уральская 6"),
    lead("Цех85, Уральская", "Пекарня-кондитерская", "Уральская улица, 6 ст1", 6, 78, "Сетевая пекарня-кондитерская рядом с производством; 2ГИС подтверждает сеть и публичные контакты бренда.", COFFEE_OFFER, "Идти только через закупки/партнерский канал, точку использовать для локального контакта.", "https://tseh85.ru/", phone="8 (800) 500-89-85", email="info@tseh85.ru", nalog_query="Цех85"),
    lead("Etlon Coffee", "Кофейня", "Средний проспект В.О., 88", 25, 79, "Кофейня внутри Baltis Plaza по карточке 2ГИС здания.", COFFEE_OFFER, "Предложить тест еды к кофе для офисного трафика БЦ.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="Etlon Coffee"),
    lead("Хлебник, Baltis Plaza", "Пекарня", "Средний проспект В.О., 88", 25, 75, "Пекарня/food-точка внутри крупного БЦ; может быть каналом допродаж или партнером.", COFFEE_OFFER, "Проверить интерес к сэндвичам и десертам вне их основной линейки.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="Хлебник Санкт-Петербург"),
    lead("Foodbrik", "Кафе/еда навынос", "Средний проспект В.О., 88", 25, 74, "Food-точка внутри Baltis Plaza; поток арендаторов офисного центра.", COFFEE_OFFER, "Обсудить поставку готовой еды как расширение ассортимента.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="Foodbrik"),
    lead("Жизньмарт, Соверен", "Супермаркет/кофе с собой", "Малый В.О. проспект, 22", 27, 80, "В 2ГИС внутри БЦ Соверен указан Жизньмарт с кофе с собой.", RETAIL_OFFER, "Проверить локального управляющего и федеральный B2B-процесс поставщика.", "https://2gis.ru/spb/firm/5348552838528578/tab/inside", nalog_query="Жизньмарт"),
    lead("Ву Ду", "Кафе", "Уральская улица, 6 ст1", 6, 68, "Кафе китайской кухни в ближайшей пешей зоне; потенциально не основной, но близкий канал.", COFFEE_OFFER, "Проверить наличие витрины и возможность кросс-продажи десертов.", nalog_query="Ву Ду кафе Уральская"),
    lead("Соседи 32", "Кафе", "улица Кораблестроителей, 32 к3", 20, 71, "Кафе в жилом массиве острова Декабристов; рядом с офисно-ритейл кластером.", COFFEE_OFFER, "Посетить точку и оценить полку/витрину для готовой еды.", nalog_query="Соседи 32 кафе"),
    lead("Булочные Ф. Вольчека, Малый В.О.", "Пекарня", "Малый В.О. проспект, 45", 24, 73, "Сетевая пекарня на маршруте к метро; потенциальный партнер или точка конкурентной разведки.", COFFEE_OFFER, "Проверить закупочный контур сети, не заходить через кассу как основной канал.", nalog_query="Булочные Ф Вольчека"),
    lead("Булочные Ф. Вольчека, Средний В.О.", "Пекарня", "Средний проспект В.О., 50", 27, 72, "Василеостровская точка сети в пределах 40 минут пешком.", COFFEE_OFFER, "Использовать как карту конкурентной полки и возможный контакт сети.", nalog_query="Булочные Ф Вольчека"),
    lead("Тарта", "Пекарня/кондитерская", "Малый В.О. проспект, 52", 25, 70, "Локальная пекарня/кондитерская в плотной пешей зоне.", COFFEE_OFFER, "Оценить ассортимент и предложить сытные позиции, если нет кухни.", nalog_query="Тарта пекарня Санкт-Петербург"),
    lead("Василеостровский рынок", "Фуд-холл/ритейл", "Большой проспект В.О., 16", 39, 73, "Фуд-кластер на границе пешего радиуса; полезен как место для партнерств и тестовых точек.", RETAIL_OFFER, "Собрать конкретных арендаторов и выбрать 3-5 food-точек для дегустации.", nalog_query="Василеостровский рынок"),
    lead("Cake Me Tender", "Кондитерская", "Большой проспект В.О., 16", 39, 65, "Точка внутри Василеостровского рынка; релевантна для десертной матрицы/партнерства.", COFFEE_OFFER, "Проверить, есть ли спрос на готовые завтраки/сэндвичи рядом с десертами.", nalog_query="Cake Me Tender"),
    lead("Забыла сахар?", "Кондитерская/кофейня", "Большой проспект В.О., 16", 39, 66, "Food-точка фуд-холла; может быть полезна для партнерских наборов.", COFFEE_OFFER, "Использовать для точечного предложения десерты+сэндвичи в одном чеке.", nalog_query="Забыла сахар"),
    lead("Супер Лента, Уральская", "Супермаркет", "Уральская улица, 4", 6, 84, "Супермаркет в ближайшей зоне; 2ГИС указывает сеть Лента и ОГРН рекламодателя.", RETAIL_OFFER, "Идти через закупки Ленты; локально проверить категорийного менеджера fresh.", "https://2gis.ru/spb/search/%D0%A3%D1%80%D0%B0%D0%BB%D1%8C%D1%81%D0%BA%D0%B0%D1%8F%204%20%D0%BB%D0%B5%D0%BD%D1%82%D0%B0", nalog_query="1037832048605", expected_ogrn="1037832048605"),
    lead("Гипер Лента, Уральская", "Гипермаркет", "Уральская улица, 29 к1", 15, 87, "Крупный гипермаркет с пекарней и кофе с собой; 2ГИС указывает 660 оценок и ОГРН ООО Лента.", RETAIL_OFFER, "Приоритетно: подготовить КП для локальной fresh-полки и федерального закупочного контура.", "https://2gis.ru/spb/firm/70000001058978791/tab/info", nalog_query="1037832048605", expected_ogrn="1037832048605"),
    lead("Магнит, Уральская", "Магазин у дома", "Уральская улица, 21 лит А", 8, 73, "Магазин в ближайшем жилом кластере; подходит для проверки спроса на ready-to-eat.", RETAIL_OFFER, "Проверить наличие fresh-полки и маршрут выхода на категорийного менеджера.", nalog_query="Магнит Уральская 21"),
    lead("РосАл", "Продуктовый магазин", "Уральская улица, 21 лит А", 8, 70, "Локальный продуктовый магазин в ЖК; проще договориться о тесте, чем с федеральной сетью.", RETAIL_OFFER, "Лично зайти с коммерческим листом и минимальной матрицей 8 SKU.", nalog_query="РосАл Уральская"),
    lead("ВкусВилл, Кораблестроителей 32/1", "Магазин полезных продуктов", "улица Кораблестроителей, 32/1", 20, 74, "2ГИС подтверждает точку ВкусВилл и публичный email бренда.", RETAIL_OFFER, "Писать в B2B/поставщики; локальную точку использовать для проверки полки.", "https://2gis.ru/spb/firm/70000001036895519", website="https://vkusvill.ru", email="info@izbenka.msk.ru", nalog_query="ВкусВилл"),
    lead("Пятёрочка, Кораблестроителей", "Магазин у дома", "улица Кораблестроителей, 30", 18, 68, "Федеральная сеть в плотной жилой зоне; потенциальный канал, но высокий барьер закупок.", RETAIL_OFFER, "Оставить как B-priority: идти через федеральный закупочный контур.", nalog_query="Пятерочка Кораблестроителей 30"),
    lead("Перекрёсток, Кораблестроителей 21", "Супермаркет", "улица Кораблестроителей, 21 к1", 18, 70, "Супермаркет в пределах 20 минут; интересен для готовой еды и перекусов.", RETAIL_OFFER, "Проверить категорию fresh и федеральный путь поставщика.", nalog_query="Перекресток Кораблестроителей 21"),
    lead("Перекрёсток, Кораблестроителей 31", "Супермаркет", "улица Кораблестроителей, 31 к1", 21, 69, "Вторая точка сети на острове Декабристов; использовать после проверки первого контакта.", RETAIL_OFFER, "Сравнить полку с первой точкой и идти через единый закупочный контур.", nalog_query="Перекресток Кораблестроителей 31"),
    lead("Азбука вкуса, Нахимова", "Премиальный супермаркет", "улица Нахимова, 5 к1", 19, 76, "Премиальная аудитория рядом с Приморской; продуктовый fit выше среднего.", RETAIL_OFFER, "Проверить требования к поставщикам и готовить premium SKU-лист.", nalog_query="Азбука вкуса Нахимова 5"),
    lead("Семишагофф, Средний В.О.", "Дискаунтер/магазин", "Средний проспект В.О., 61", 26, 70, "Локальная сеть СПб/ЛО; может быть проще федеральных сетей для переговоров.", RETAIL_OFFER, "Проверить закупочный email и предложить доступную линейку перекусов.", phone="+7 (812) 648-50-00", email="sklad.logist1@7shagov.org", nalog_query="Семишагофф"),
    lead("Семишагофф, Одоевского", "Дискаунтер/магазин", "улица Одоевского, 27", 17, 72, "Точка сети ближе к производству и плотному жилому потоку.", RETAIL_OFFER, "Идти через сеть, локально проверить ассортимент ready-to-eat.", phone="+7 (812) 648-50-00", email="sklad.logist1@7shagov.org", nalog_query="Семишагофф"),
    lead("ТЦ Пальмира", "Торговый центр", "улица Кораблестроителей, 48 к2", 31, 74, "ТЦ у жилого массива; интересен как кластер арендаторов и мини-витрин.", RETAIL_OFFER, "Собрать арендаторов food/retail и предложить поставку в один адрес.", nalog_query="Пальмира торговый центр Кораблестроителей 48"),
    lead("Шкиперский Молл", "Торговый центр", "Малый В.О. проспект, 88", 36, 72, "ТЦ на границе 40-минутной зоны; подходит для обхода арендаторов.", RETAIL_OFFER, "Проверить арендаторов и food-точки, затем сделать маршрутный визит.", nalog_query="Шкиперский Молл"),
    lead("Морской завод Алмаз", "Производство/офис", "Уральская улица, 19", 7, 82, "Крупный производственный/офисный резидент рядом; потенциальный корпоративный спрос на перекусы.", OFFICE_OFFER, "Предложить корпоративную поставку в офис/производственный блок.", "https://2gis.ru/spb/firm/70000001045044337/tab/inside", nalog_query="Морской завод Алмаз"),
    lead("Яндекс Технологии", "Офисный арендатор", "Уральская улица, 1 к2", 10, 84, "Офисный арендатор в БЦ Урок; высокий потенциал микромаркета для сотрудников.", OFFICE_OFFER, "Искать office/facility manager через БЦ или публичный корпоративный канал.", "https://2gis.ru/spb/firm/5348552838484493/tab/inside", nalog_query="Яндекс Технологии"),
    lead("Skyeng", "Офисный арендатор", "Уральская улица, 1 к2", 10, 78, "Офисная/образовательная компания в БЦ Урок; интересна для корпоративных перекусов.", OFFICE_OFFER, "Проверить локальный офис и предложить наборы на команды.", "https://2gis.ru/spb/firm/5348552838484493/tab/inside", nalog_query="Skyeng"),
    lead("Бланк банк", "Офисный арендатор", "Уральская улица, 1 к2", 10, 76, "Офисный арендатор в БЦ Урок; может быть потребность в перекусах для сотрудников.", OFFICE_OFFER, "Выход через офис-менеджера/администрацию БЦ.", "https://2gis.ru/spb/firm/5348552838484493/tab/inside", nalog_query="Бланк банк"),
    lead("ПлазаТелеком", "Офисный арендатор", "Средний проспект В.О., 88", 25, 73, "Офисный арендатор Baltis Plaza; подходит для корпоративной поставки.", OFFICE_OFFER, "Проверить размер офиса и предложить еженедельную витрину.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="ПлазаТелеком"),
    lead("Росконгресс", "Офисный арендатор", "Средний проспект В.О., 88", 25, 78, "Организационный/событийный профиль: возможны офисные и event-поставки.", OFFICE_OFFER, "Искать административный контакт и предложить наборы для мероприятий.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="Росконгресс"),
    lead("Мать и дитя, Baltis Plaza", "Медицинский центр", "Средний проспект В.О., 88", 25, 77, "Медицинский арендатор БЦ; спрос на аккуратные перекусы для персонала/посетителей.", MEDICAL_OFFER, "Предложить отдельную матрицу без резких запахов и с понятными сроками.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="Мать и дитя Средний 88"),
    lead("Dental Story", "Клиника", "Средний проспект В.О., 88", 25, 69, "Клиника внутри БЦ; потенциально небольшой корпоративный спрос.", MEDICAL_OFFER, "Проверить численность команды и формат комнаты отдыха.", "https://2gis.ru/spb/firm/5348552838657634/tab/inside", nalog_query="Dental Story"),
    lead("Райффайзенбанк, Соверен", "Банк/офис", "Малый В.О. проспект, 22", 27, 68, "Банк на первом этаже БЦ Соверен; потенциально только офисная поставка.", OFFICE_OFFER, "Низкий приоритет: использовать как часть предложения для всего БЦ.", "https://2gis.ru/spb/firm/5348552838528578/tab/inside", nalog_query="Райффайзенбанк"),
    lead("Банк ВТБ, Соверен", "Банк/офис", "Малый В.О. проспект, 22", 27, 68, "Отделение в БЦ Соверен; лучше заходить через управляющую компанию здания.", OFFICE_OFFER, "Включить в пакет арендаторов БЦ, не отдельный первый контакт.", "https://2gis.ru/spb/firm/5348552838528578/tab/inside", nalog_query="Банк ВТБ"),
    lead("Business FM", "Офис/медиа", "Малый В.О. проспект, 22", 27, 73, "Офис в БЦ Соверен; команда и гости могут потреблять перекусы.", OFFICE_OFFER, "Предложить корпоративные наборы на редакционные смены/гостей.", "https://2gis.ru/spb/firm/5348552838528578/tab/inside", nalog_query="Business FM Санкт-Петербург"),
    lead("Wellmart24", "Торговая компания", "Малый В.О. проспект, 22", 27, 72, "Торговая компания в БЦ Соверен; может быть партнерским B2B-каналом.", OFFICE_OFFER, "Проверить профиль торговли и возможность партнерского размещения.", "https://2gis.ru/spb/firm/5348552838528578/tab/inside", nalog_query="Wellmart24"),
    lead("Диалог", "Коворкинг", "набережная реки Смоленки, 3 к1", 9, 85, "Коворкинг/офисный формат очень близко к производству; высокий fit для регулярной полки.", COWORKING_OFFER, "Зайти лично с дегустацией и предложить холодильник на 2 недели.", nalog_query="Диалог коворкинг Смоленки 3"),
    lead("База", "Коворкинг", "12-я линия В.О., 13", 35, 80, "Коворкинг на Васильевском острове; высокий fit для перекусов резидентов.", COWORKING_OFFER, "Написать администратору и предложить пилот с учетом списаний.", nalog_query="База коворкинг 12 линия"),
    lead("Стрижи", "Коворкинг", "Кадетская линия В.О., 27/5", 37, 76, "Коворкинг/офисная площадка на границе радиуса.", COWORKING_OFFER, "Проверить резидентский поток и предложить компактную матрицу.", nalog_query="Стрижи коворкинг"),
    lead("Lady Lu", "Коворкинг/пространство", "4-я линия В.О., 41А", 36, 69, "Небольшое пространство/коворкинг; может быть точкой для десертов и сэндвичей.", COWORKING_OFFER, "Проверить фактический формат и посещаемость.", nalog_query="Lady Lu коворкинг"),
    lead("Docklands", "Апарт-отель", "набережная Макарова, 60", 20, 82, "Апарт-отель и жилой кластер рядом с производством; подходит для завтраков и лобби-витрины.", HOTEL_OFFER, "Выйти на управляющую компанию/службу размещения с тестовым снэк-боксом.", nalog_query="Docklands апарт отель"),
    lead("Marco Polo", "Отель", "12-я линия В.О., 27", 35, 73, "Отель в пределах 40 минут; возможны завтраки, ланч-боксы, мини-бар.", HOTEL_OFFER, "Предложить B2B-прайс для завтраков/групповых заездов.", nalog_query="Marco Polo отель Санкт-Петербург"),
    lead("Саквояж", "Отель", "Средний проспект В.О., 48/27", 30, 72, "Отель на Среднем проспекте В.О.; подходит для малых регулярных поставок.", HOTEL_OFFER, "Проверить текущий формат завтраков и предложить готовые позиции.", nalog_query="Саквояж отель Средний проспект"),
    lead("Olympia Medical Center", "Медицинский центр", "Средний проспект В.О., 38 к1", 30, 70, "Медицинский арендатор/сосед в БЦ Gustaf; потенциальная витрина для персонала.", MEDICAL_OFFER, "Предложить небольшой корпоративный набор, не как розничную полку.", "https://2gis.ru/spb/geo/5348720342208864", nalog_query="Olympia Medical Center"),
]


def fns_lookup(query: str, expected_ogrn: str | None, cache: dict) -> dict:
    if not query:
        return {"status": "not_checked"}
    if not expected_ogrn:
        return {"status": "manual_pb_match", "total": None}
    key = f"{query}|{expected_ogrn or ''}"
    if key in cache:
        return cache[key]
    try:
        data = urllib.parse.urlencode({"query": query, "region": "", "PreventChromeAutocomplete": ""}).encode("utf-8")
        req = urllib.request.Request(
            "https://egrul.nalog.ru/",
            data=data,
            headers={"User-Agent": "Mozilla/5.0", "Content-Type": "application/x-www-form-urlencoded"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=20) as response:
            token = json.loads(response.read().decode("utf-8")).get("t")
        time.sleep(0.6)
        req2 = urllib.request.Request(f"https://egrul.nalog.ru/search-result/{token}", headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req2, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
        rows = payload.get("rows") or []
        match = None
        if expected_ogrn:
            match = next((row for row in rows if str(row.get("o")) == str(expected_ogrn)), None)
        if not match and rows:
            active_rows = [row for row in rows if not row.get("e")]
            match = active_rows[0] if active_rows else rows[0]
        if not match:
            result = {"status": "no_match", "total": 0}
        else:
            result = {
                "status": "verified_ogrn" if expected_ogrn and str(match.get("o")) == str(expected_ogrn) else "candidate_found",
                "total": len(rows),
                "legal_name": match.get("n") or match.get("c"),
                "inn": match.get("i"),
                "ogrn": match.get("o"),
                "kpp": match.get("p"),
                "region": match.get("rn"),
                "manager": match.get("g"),
                "closed_date": match.get("e"),
            }
    except Exception as error:  # noqa: BLE001
        result = {"status": "lookup_error", "error": str(error)}
    cache[key] = result
    return result


def enrich_with_fns(rows: list[dict]) -> list[dict]:
    LOG_DIR.mkdir(exist_ok=True)
    cache_path = LOG_DIR / "uralskaya_fns_cache.json"
    cache = json.loads(cache_path.read_text("utf-8")) if cache_path.exists() else {}
    enriched = []
    for row in rows:
        fns = fns_lookup(row["nalog_query"], row.get("expected_ogrn"), cache)
        enriched_row = {**row}
        enriched_row.update(
            {
                "fns_status": fns.get("status"),
                "legal_name": fns.get("legal_name"),
                "inn": fns.get("inn"),
                "ogrn": fns.get("ogrn"),
                "kpp": fns.get("kpp"),
                "fns_region": fns.get("region"),
                "fns_total_matches": fns.get("total"),
                "fns_notes": fns.get("error") or ("Точное совпадение по ОГРН" if fns.get("status") == "verified_ogrn" else "Кандидат из ЕГРЮЛ; нужен ручной матч с точкой" if fns.get("status") == "candidate_found" else "Не найдено/не проверено"),
            }
        )
        enriched.append(enriched_row)
        time.sleep(0.2)
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), "utf-8")
    return enriched


HEADERS = [
    ("name", "Компания/точка"),
    ("segment", "Сегмент"),
    ("address", "Адрес"),
    ("walk_min", "Пешком, мин"),
    ("distance_band", "Зона"),
    ("priority", "Приоритет"),
    ("score", "Score"),
    ("fit_reason", "Почему подходит"),
    ("offer", "Что предложить"),
    ("next_action", "Следующее действие"),
    ("phone", "Телефон"),
    ("email", "Email"),
    ("website", "Сайт"),
    ("source_2gis", "2ГИС"),
    ("source_yandex", "Яндекс Карты"),
    ("pb_nalog_url", "Прозрачный бизнес"),
    ("nalog_query", "Запрос ФНС"),
    ("fns_status", "Статус ФНС"),
    ("legal_name", "Юрлицо ФНС"),
    ("inn", "ИНН"),
    ("ogrn", "ОГРН"),
    ("fns_notes", "Комментарий проверки"),
    ("notes", "Примечания"),
]


def write_json(rows: list[dict]) -> Path:
    DATA_DIR.mkdir(exist_ok=True)
    path = DATA_DIR / "uralskaya-local-prospects.json"
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "production_address": PRODUCTION_ADDRESS,
        "walk_radius_min": WALK_RADIUS_MIN,
        "walk_radius_km": WALK_RADIUS_KM,
        "method": "Открытые карточки 2ГИС/Яндекс Карты, ручная sales-квалификация и проверочный слой ФНС ЕГРЮЛ/Прозрачный бизнес.",
        "rows": rows,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), "utf-8")
    return path


def write_sqlite(rows: list[dict]) -> None:
    if not DB_PATH.exists():
        return
    con = sqlite3.connect(DB_PATH)
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS local_prospects (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          name TEXT NOT NULL,
          segment TEXT NOT NULL,
          address TEXT NOT NULL,
          walk_min INTEGER,
          distance_band TEXT,
          priority TEXT,
          score INTEGER,
          fit_reason TEXT,
          offer TEXT,
          next_action TEXT,
          phone TEXT,
          email TEXT,
          website TEXT,
          source_2gis TEXT,
          source_yandex TEXT,
          pb_nalog_url TEXT,
          nalog_query TEXT,
          fns_status TEXT,
          legal_name TEXT,
          inn TEXT,
          ogrn TEXT,
          fns_notes TEXT,
          notes TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    con.execute("DELETE FROM local_prospects")
    fields = [key for key, _ in HEADERS]
    placeholders = ",".join("?" for _ in fields)
    con.executemany(
        f"INSERT INTO local_prospects ({','.join(fields)}) VALUES ({placeholders})",
        [[row.get(field) for field in fields] for row in rows],
    )
    con.commit()
    con.close()


def write_xlsx(rows: list[dict]) -> Path:
    OUT_DIR.mkdir(exist_ok=True)
    path = OUT_DIR / "uralskaya_40min_prospects_2026-06-03.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "SPB+LO seed"
    ws.append([label for _, label in HEADERS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(key) for key, _ in HEADERS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 28,
        "B": 18,
        "C": 34,
        "D": 12,
        "E": 12,
        "F": 10,
        "G": 8,
        "H": 42,
        "I": 44,
        "J": 42,
        "N": 34,
        "O": 34,
        "P": 34,
        "S": 36,
        "V": 34,
        "W": 36,
    }
    for idx in range(1, len(HEADERS) + 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, 18)
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(2, ws.max_row + 1):
        priority = ws.cell(row=row_idx, column=6).value
        fill = "D9EAD3" if priority == "A" else "FFF2CC" if priority == "B" else "FCE4D6"
        ws.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor=fill)
    summary = wb.create_sheet("Сводка")
    summary["A1"] = "Локальная seed-база Lunch Up для SPB+ЛО outreach"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Радиус: {WALK_RADIUS_MIN} минут пешком, ориентир {WALK_RADIUS_KM} км"
    summary["A3"] = f"Дата сборки: {TODAY}"
    segment_counts = Counter(row["segment"] for row in rows)
    priority_counts = Counter(row["priority"] for row in rows)
    summary.append([])
    summary.append(["Сегмент", "Количество"])
    for segment, count in segment_counts.most_common():
        summary.append([segment, count])
    start = summary.max_row + 2
    summary.cell(start, 1, "Приоритет")
    summary.cell(start, 2, "Количество")
    for idx, priority in enumerate(["A", "B", "C"], start=start + 1):
        summary.cell(idx, 1, priority)
        summary.cell(idx, 2, priority_counts.get(priority, 0))
    notes_start = summary.max_row + 2
    notes = [
        "Метод: открытые карточки 2ГИС/Яндекс Карты, затем sales-score по близости, сегменту и простоте пилота.",
        "ФНС: поле 'Статус ФНС' показывает машинную проверку через официальный ЕГРЮЛ; 'candidate_found' требует ручного матча с конкретной точкой.",
        "Не использовать как холодную массовую рассылку. Для сетей идти через поставщиков/закупки, для БЦ и коворкингов — через администрацию.",
    ]
    for offset, note in enumerate(notes):
        summary.cell(notes_start + offset, 1, note)
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 48 if col == 1 else 16
    wb.save(path)
    return path


def write_html(rows: list[dict]) -> Path:
    OUT_DIR.mkdir(exist_ok=True)
    path = OUT_DIR / "uralskaya_40min_prospects_2026-06-03.html"
    segment_counts = Counter(row["segment"] for row in rows)
    priority_counts = Counter(row["priority"] for row in rows)
    top_rows = sorted(rows, key=lambda item: (-item["score"], item["walk_min"], item["name"]))
    table_rows = []
    for row in top_rows:
        links = [
            f'<a href="{escape(row["source_2gis"])}" target="_blank">2ГИС</a>',
            f'<a href="{escape(row["source_yandex"])}" target="_blank">Яндекс</a>',
            f'<a href="{escape(row["pb_nalog_url"])}" target="_blank">ФНС</a>',
        ]
        contact = "<br>".join(filter(None, [escape(row.get("phone") or ""), escape(row.get("email") or "")])) or "через карту/сайт"
        table_rows.append(
            "<tr>"
            f"<td><b>{escape(row['name'])}</b><div class='muted'>{escape(row['address'])}</div></td>"
            f"<td>{escape(row['segment'])}</td>"
            f"<td>{row['walk_min']} мин<br><span class='pill'>{escape(row['distance_band'])}</span></td>"
            f"<td><span class='prio prio-{escape(row['priority'])}'>{escape(row['priority'])}</span><br>{row['score']}</td>"
            f"<td>{escape(row['fit_reason'])}</td>"
            f"<td>{escape(row['offer'])}<div class='muted'>{escape(row['next_action'])}</div></td>"
            f"<td>{contact}</td>"
            f"<td>{escape(row.get('fns_status') or '')}<div class='muted'>{escape(row.get('inn') or '')}</div></td>"
            f"<td>{' / '.join(links)}</td>"
            "</tr>"
        )
    segment_bars = "\n".join(
        f"<div class='bar'><span>{escape(segment)}</span><b style='width:{count / max(segment_counts.values()) * 100:.0f}%'></b><em>{count}</em></div>"
        for segment, count in segment_counts.most_common()
    )
    priority_cards = "\n".join(
        f"<div class='card'><div class='label'>Приоритет {priority}</div><div class='value'>{priority_counts.get(priority, 0)}</div></div>"
        for priority in ["A", "B", "C"]
    )
    html = f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Lunch Up: локальная seed-база SPB+ЛО</title>
  <style>
    body{{margin:0;background:#f6f8fb;color:#152033;font:14px/1.45 Arial,Helvetica,sans-serif}}
    main{{max-width:1480px;margin:0 auto;padding:24px}}
    h1{{font-size:28px;margin:0 0 8px}} h2{{margin-top:26px}} p{{color:#5b6778}}
    .hero,.card,.panel{{background:#fff;border:1px solid #dce3ee;border-radius:8px;padding:18px;box-shadow:0 1px 2px #0f172a0f}}
    .grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin:16px 0}}
    .label{{font-size:12px;color:#667085;text-transform:uppercase;font-weight:700}} .value{{font-size:26px;font-weight:800;margin-top:4px}}
    table{{width:100%;border-collapse:collapse;background:#fff}} th,td{{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left;vertical-align:top}} th{{font-size:11px;text-transform:uppercase;color:#64748b;background:#f8fafc;position:sticky;top:0}}
    .muted{{color:#64748b;font-size:12px;margin-top:4px}} a{{color:#155e9f;text-decoration:none}} .pill{{display:inline-block;border:1px solid #d7dfeb;border-radius:6px;padding:2px 6px;color:#475569;font-size:12px}}
    .prio{{display:inline-block;border-radius:6px;padding:4px 8px;font-weight:800}} .prio-A{{background:#dcfce7;color:#166534}} .prio-B{{background:#fef3c7;color:#92400e}} .prio-C{{background:#fee2e2;color:#991b1b}}
    .bar{{display:grid;grid-template-columns:210px 1fr 40px;align-items:center;gap:10px;margin:8px 0}} .bar b{{display:block;background:#1d4f8f;height:12px;border-radius:4px}} .bar em{{font-style:normal;color:#475569}}
    .toolbar{{display:flex;gap:8px;flex-wrap:wrap;margin:14px 0}} input,select{{height:36px;border:1px solid #cfd8e3;border-radius:6px;padding:0 10px;background:#fff}}
    @media(max-width:900px){{.grid{{grid-template-columns:1fr}} table{{display:block;overflow:auto;white-space:nowrap}}}}
  </style>
</head>
<body>
<main>
  <section class="hero">
    <div class="label">Lunch Up / локальная база продаж</div>
    <h1>Локальная seed-база для SPB+ЛО outreach</h1>
    <p>База собрана по открытым карточкам 2ГИС/Яндекс Карт и проверочному слою ФНС. Пеший радиус от производства сохранен как один из источников первого пилота, но актуальная стратегия масштабируется на СПб и Ленинградскую область через согласованные маршруты.</p>
  </section>
  <section class="grid">
    <div class="card"><div class="label">Всего лидов</div><div class="value">{len(rows)}</div></div>
    <div class="card"><div class="label">Приоритет A</div><div class="value">{priority_counts.get('A', 0)}</div></div>
    <div class="card"><div class="label">До 10 минут</div><div class="value">{sum(1 for row in rows if row['walk_min'] <= 10)}</div></div>
    <div class="card"><div class="label">Средний score</div><div class="value">{round(sum(row['score'] for row in rows) / len(rows))}</div></div>
  </section>
  <section class="panel">
    <h2>Сегменты</h2>
    {segment_bars}
  </section>
  <section class="grid">{priority_cards}</section>
  <section class="panel">
    <h2>База лидов</h2>
    <div class="toolbar">
      <input id="q" placeholder="Поиск по названию, адресу, сегменту">
      <select id="priority"><option value="">Все приоритеты</option><option>A</option><option>B</option><option>C</option></select>
    </div>
    <table id="leads">
      <thead><tr><th>Компания</th><th>Сегмент</th><th>Пешком</th><th>Приоритет</th><th>Почему подходит</th><th>Оффер</th><th>Контакт</th><th>ФНС</th><th>Источники</th></tr></thead>
      <tbody>{''.join(table_rows)}</tbody>
    </table>
  </section>
</main>
<script>
const q=document.getElementById('q'), p=document.getElementById('priority'), rows=[...document.querySelectorAll('#leads tbody tr')];
function filter(){{const needle=q.value.toLowerCase().trim(), pr=p.value; rows.forEach(r=>{{const ok=(!needle||r.textContent.toLowerCase().includes(needle))&&(!pr||r.querySelector('.prio')?.textContent===pr); r.style.display=ok?'':'none';}})}}
q.addEventListener('input',filter); p.addEventListener('change',filter);
</script>
</body>
</html>"""
    path.write_text(html, "utf-8")
    return path


def main() -> None:
    rows = enrich_with_fns(PROSPECTS)
    json_path = write_json(rows)
    write_sqlite(rows)
    xlsx_path = write_xlsx(rows)
    html_path = write_html(rows)
    summary = {
        "rows": len(rows),
        "priority": Counter(row["priority"] for row in rows),
        "segments": Counter(row["segment"] for row in rows),
        "json": str(json_path),
        "xlsx": str(xlsx_path),
        "html": str(html_path),
        "sqlite_table": "local_prospects",
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=dict))


if __name__ == "__main__":
    main()
